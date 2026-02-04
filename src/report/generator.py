"""
Report generation utilities for PDF and Excel formats.
"""
import os
import io
from datetime import datetime, date
from typing import List, Dict, Any, Optional
from dataclasses import dataclass
from loguru import logger

# PDF generation
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch, cm
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
    PageBreak, Image, ListFlowable, ListItem
)
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

# Excel generation
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, PieChart, Reference


@dataclass
class ReportConfig:
    """Report configuration."""
    title: str
    subtitle: Optional[str] = None
    author: str = "Signal Transceiver"
    date_range: Optional[tuple] = None
    logo_path: Optional[str] = None
    include_summary: bool = True
    include_charts: bool = True
    page_size: str = "A4"


class PDFReportGenerator:
    """Generate PDF reports."""

    def __init__(self, config: ReportConfig):
        self.config = config
        self.styles = getSampleStyleSheet()
        self._setup_custom_styles()

    def _setup_custom_styles(self):
        """Setup custom paragraph styles."""
        self.styles.add(ParagraphStyle(
            name='ReportTitle',
            parent=self.styles['Heading1'],
            fontSize=24,
            spaceAfter=30,
            alignment=TA_CENTER
        ))
        self.styles.add(ParagraphStyle(
            name='ReportSubtitle',
            parent=self.styles['Normal'],
            fontSize=14,
            spaceAfter=20,
            alignment=TA_CENTER,
            textColor=colors.grey
        ))
        self.styles.add(ParagraphStyle(
            name='SectionTitle',
            parent=self.styles['Heading2'],
            fontSize=16,
            spaceBefore=20,
            spaceAfter=10,
            textColor=colors.darkblue
        ))
        self.styles.add(ParagraphStyle(
            name='TableHeader',
            parent=self.styles['Normal'],
            fontSize=10,
            textColor=colors.white,
            alignment=TA_CENTER
        ))

    def generate(
        self,
        data: Dict[str, Any],
        output_path: Optional[str] = None
    ) -> bytes:
        """Generate PDF report."""
        buffer = io.BytesIO()

        page_size = A4 if self.config.page_size == "A4" else letter
        doc = SimpleDocTemplate(
            buffer,
            pagesize=page_size,
            rightMargin=72,
            leftMargin=72,
            topMargin=72,
            bottomMargin=72
        )

        story = []

        # Title page
        story.extend(self._build_title_page())

        # Summary section
        if self.config.include_summary and "summary" in data:
            story.extend(self._build_summary_section(data["summary"]))

        # Data tables
        if "tables" in data:
            for table_config in data["tables"]:
                story.extend(self._build_table_section(table_config))

        # Additional sections
        if "sections" in data:
            for section in data["sections"]:
                story.extend(self._build_custom_section(section))

        # Build PDF
        doc.build(story)
        pdf_content = buffer.getvalue()
        buffer.close()

        # Save to file if path provided
        if output_path:
            os.makedirs(os.path.dirname(output_path), exist_ok=True)
            with open(output_path, 'wb') as f:
                f.write(pdf_content)
            logger.info(f"PDF report saved to: {output_path}")

        return pdf_content

    def _build_title_page(self) -> List:
        """Build title page elements."""
        elements = []
        elements.append(Spacer(1, 2 * inch))
        elements.append(Paragraph(self.config.title, self.styles['ReportTitle']))

        if self.config.subtitle:
            elements.append(Paragraph(self.config.subtitle, self.styles['ReportSubtitle']))

        elements.append(Spacer(1, inch))

        # Date info
        date_str = datetime.now().strftime("%Y年%m月%d日")
        if self.config.date_range:
            start, end = self.config.date_range
            date_str = f"{start} 至 {end}"

        elements.append(Paragraph(
            f"报告日期: {date_str}",
            self.styles['Normal']
        ))
        elements.append(Paragraph(
            f"生成者: {self.config.author}",
            self.styles['Normal']
        ))
        elements.append(PageBreak())

        return elements

    def _build_summary_section(self, summary: Dict[str, Any]) -> List:
        """Build summary section."""
        elements = []
        elements.append(Paragraph("概要", self.styles['SectionTitle']))

        if isinstance(summary, dict):
            items = []
            for key, value in summary.items():
                items.append(ListItem(Paragraph(
                    f"<b>{key}:</b> {value}",
                    self.styles['Normal']
                )))
            elements.append(ListFlowable(items, bulletType='bullet'))
        else:
            elements.append(Paragraph(str(summary), self.styles['Normal']))

        elements.append(Spacer(1, 0.5 * inch))
        return elements

    def _build_table_section(self, table_config: Dict[str, Any]) -> List:
        """Build table section."""
        elements = []

        title = table_config.get("title", "数据表")
        headers = table_config.get("headers", [])
        rows = table_config.get("rows", [])

        elements.append(Paragraph(title, self.styles['SectionTitle']))

        if not rows:
            elements.append(Paragraph("暂无数据", self.styles['Normal']))
            return elements

        # Build table data
        table_data = [headers] + rows

        # Create table
        col_widths = table_config.get("col_widths", None)
        table = Table(table_data, colWidths=col_widths)

        # Style table
        style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
        ])
        table.setStyle(style)

        elements.append(table)
        elements.append(Spacer(1, 0.5 * inch))

        return elements

    def _build_custom_section(self, section: Dict[str, Any]) -> List:
        """Build custom section."""
        elements = []

        title = section.get("title", "")
        content = section.get("content", "")

        if title:
            elements.append(Paragraph(title, self.styles['SectionTitle']))

        if content:
            elements.append(Paragraph(content, self.styles['Normal']))

        elements.append(Spacer(1, 0.3 * inch))
        return elements


class ExcelReportGenerator:
    """Generate Excel reports."""

    def __init__(self, config: ReportConfig):
        self.config = config
        self.workbook = None
        self._header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        self._header_font = Font(color="FFFFFF", bold=True)
        self._border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def generate(
        self,
        data: Dict[str, Any],
        output_path: Optional[str] = None
    ) -> bytes:
        """Generate Excel report."""
        self.workbook = Workbook()

        # Remove default sheet
        if 'Sheet' in self.workbook.sheetnames:
            del self.workbook['Sheet']

        # Summary sheet
        if self.config.include_summary and "summary" in data:
            self._create_summary_sheet(data["summary"])

        # Data sheets
        if "sheets" in data:
            for sheet_config in data["sheets"]:
                self._create_data_sheet(sheet_config)

        # Save to buffer
        buffer = io.BytesIO()
        self.workbook.save(buffer)
        excel_content = buffer.getvalue()
        buffer.close()

        # Save to file if path provided
        if output_path:
            os.makedirs(os.path.dirname(output_path), exist_ok=True)
            with open(output_path, 'wb') as f:
                f.write(excel_content)
            logger.info(f"Excel report saved to: {output_path}")

        return excel_content

    def _create_summary_sheet(self, summary: Dict[str, Any]):
        """Create summary sheet."""
        ws = self.workbook.create_sheet("概要", 0)

        # Title
        ws.merge_cells('A1:D1')
        ws['A1'] = self.config.title
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')

        # Subtitle
        if self.config.subtitle:
            ws.merge_cells('A2:D2')
            ws['A2'] = self.config.subtitle
            ws['A2'].alignment = Alignment(horizontal='center')

        # Summary data
        row = 4
        for key, value in summary.items():
            ws.cell(row=row, column=1, value=key).font = Font(bold=True)
            ws.cell(row=row, column=2, value=str(value))
            row += 1

        # Auto-adjust column widths
        self._auto_adjust_columns(ws)

    def _create_data_sheet(self, sheet_config: Dict[str, Any]):
        """Create data sheet."""
        name = sheet_config.get("name", "数据")[:31]  # Excel limit
        ws = self.workbook.create_sheet(name)

        headers = sheet_config.get("headers", [])
        rows = sheet_config.get("rows", [])

        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = self._header_fill
            cell.font = self._header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = self._border

        # Write data rows
        for row_idx, row_data in enumerate(rows, 2):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = self._border
                cell.alignment = Alignment(horizontal='center')

        # Add chart if configured
        if self.config.include_charts and sheet_config.get("chart"):
            self._add_chart(ws, sheet_config["chart"], len(rows))

        # Auto-adjust column widths
        self._auto_adjust_columns(ws)

    def _add_chart(self, ws, chart_config: Dict[str, Any], data_rows: int):
        """Add chart to worksheet."""
        chart_type = chart_config.get("type", "bar")
        title = chart_config.get("title", "图表")

        if chart_type == "bar":
            chart = BarChart()
        elif chart_type == "line":
            chart = LineChart()
        elif chart_type == "pie":
            chart = PieChart()
        else:
            return

        chart.title = title

        data_col = chart_config.get("data_column", 2)
        label_col = chart_config.get("label_column", 1)

        data = Reference(ws, min_col=data_col, min_row=1, max_row=data_rows + 1)
        labels = Reference(ws, min_col=label_col, min_row=2, max_row=data_rows + 1)

        chart.add_data(data, titles_from_data=True)
        chart.set_categories(labels)

        ws.add_chart(chart, chart_config.get("position", "E2"))

    def _auto_adjust_columns(self, ws):
        """Auto-adjust column widths based on content."""
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width


class ReportService:
    """Service for generating reports."""

    def __init__(self):
        self.output_dir = "reports"
        os.makedirs(self.output_dir, exist_ok=True)

    async def generate_data_report(
        self,
        data_records: List[Dict[str, Any]],
        format: str = "pdf",
        title: str = "数据报告"
    ) -> bytes:
        """Generate a data report."""
        config = ReportConfig(
            title=title,
            subtitle=f"共 {len(data_records)} 条记录",
            include_summary=True,
            include_charts=True
        )

        # Prepare data
        report_data = {
            "summary": {
                "总记录数": len(data_records),
                "报告时间": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            }
        }

        if data_records:
            headers = ["ID", "类型", "标的", "日期", "状态", "创建时间"]
            rows = [
                [
                    r.get("id", ""),
                    r.get("type", ""),
                    r.get("symbol", ""),
                    str(r.get("execute_date", "")),
                    r.get("status", ""),
                    str(r.get("created_at", ""))[:19]
                ]
                for r in data_records
            ]

            if format == "pdf":
                report_data["tables"] = [{"title": "数据记录", "headers": headers, "rows": rows}]
            else:
                report_data["sheets"] = [{"name": "数据记录", "headers": headers, "rows": rows}]

        # Generate report
        if format == "pdf":
            generator = PDFReportGenerator(config)
        else:
            generator = ExcelReportGenerator(config)

        return generator.generate(report_data)

    async def generate_performance_report(
        self,
        stats: Dict[str, Any],
        history: List[Dict[str, Any]],
        format: str = "pdf"
    ) -> bytes:
        """Generate a performance report."""
        config = ReportConfig(
            title="系统性能报告",
            subtitle="Signal Transceiver 性能监控",
            include_summary=True,
            include_charts=True
        )

        report_data = {
            "summary": {
                "运行时间": f"{stats.get('uptime_seconds', 0) / 3600:.1f} 小时",
                "总请求数": stats.get("total_requests", 0),
                "总错误数": stats.get("total_errors", 0),
                "平均响应时间": f"{stats.get('avg_response_time_ms', 0):.1f}ms",
                "CPU使用率": f"{stats.get('cpu_percent', 0):.1f}%",
                "内存使用率": f"{stats.get('memory_percent', 0):.1f}%"
            }
        }

        if history:
            headers = ["时间", "CPU%", "内存%", "请求/秒", "响应时间(ms)", "错误率"]
            rows = [
                [
                    h.get("timestamp", "")[:19],
                    f"{h.get('cpu_percent', 0):.1f}",
                    f"{h.get('memory_percent', 0):.1f}",
                    f"{h.get('requests_per_second', 0):.2f}",
                    f"{h.get('avg_response_time_ms', 0):.1f}",
                    f"{h.get('error_rate', 0) * 100:.2f}%"
                ]
                for h in history[-100:]  # Last 100 records
            ]

            if format == "pdf":
                report_data["tables"] = [{"title": "性能历史", "headers": headers, "rows": rows}]
            else:
                report_data["sheets"] = [{"name": "性能历史", "headers": headers, "rows": rows}]

        if format == "pdf":
            generator = PDFReportGenerator(config)
        else:
            generator = ExcelReportGenerator(config)

        return generator.generate(report_data)


# Global report service instance
report_service = ReportService()
